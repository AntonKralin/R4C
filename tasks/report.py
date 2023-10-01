from openpyxl import Workbook
from django.db.models import Count
from django.utils import timezone

from datetime import timedelta

from robots.models import Robot


def create_file(days: int = 7) -> str:
    filename = 'week.xlsx'
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)
    headers = ['model', 'version', 'count']

    now = timezone.now()
    week = now - timedelta(days=days)
    gp_robot = Robot.objects.filter(created__gte=week) \
                    .values('model', 'version') \
                    .annotate(total=Count('*')) \
                    .order_by('model')

    prev = None
    count = 2
    for i_dict in gp_robot:
        if prev != i_dict['model']:
            prev = i_dict['model']
            ws = wb.create_sheet(i_dict['model'])
            count = 2

            for i, name in enumerate(headers):
                cell = ws.cell(1, i+1)
                cell.value = name
        else:
            count += 1
        cell = ws.cell(count, 1)
        cell.value = i_dict['model']
        cell = ws.cell(count, 2)
        cell.value = i_dict['version']
        cell = ws.cell(count, 3)
        cell.value = i_dict['total']

    wb.save(filename=filename)
    return filename
